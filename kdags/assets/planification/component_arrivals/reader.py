import re
from datetime import datetime, timedelta
from io import BytesIO

import openpyxl
import pandas as pd
from office365.graph_client import GraphClient

from kdags.resources.msgraph import acquire_token_func


def rename_datetime_columns(df):
    # Function to convert datetime string to date string
    def to_date_string(col_name):
        try:
            return datetime.strptime(str(col_name), "%Y-%m-%d %H:%M:%S").strftime("%Y-%m-%d")
        except ValueError:
            return col_name  # Return original if not a datetime string

    # Create a dictionary of old_name: new_name for columns
    rename_dict = {col: to_date_string(col) for col in df.columns}

    # Rename the columns
    df.rename(columns=rename_dict, inplace=True)

    return df


def reshape_to_long_format(df):
    # Identify date columns
    date_columns = [col for col in df.columns if pd.to_datetime(col, errors="coerce") is not pd.NaT]

    # Identify non-date columns
    id_vars = [col for col in df.columns if col not in date_columns]

    # Melt the DataFrame
    df_long = df.melt(
        id_vars=id_vars,
        value_vars=date_columns,
        var_name="arrival_week",
        value_name="value",
    )

    # Convert arrival_date to datetime
    df_long["arrival_week"] = df_long["arrival_week"].apply(convert_to_iso_week)

    # Sort the DataFrame
    df_long = df_long.sort_values(by=["Componente", "arrival_week"])

    return df_long


def convert_to_iso_week(date_str):
    date = pd.to_datetime(date_str)
    if date.weekday() != 0:
        raise ValueError(f"Date {date_str} is not a Monday")
    return date.strftime("%Y-W%V")


def extract_date_and_type(value):
    if pd.isna(value):
        return pd.NaT, "PROYECTADO"

    value_str = str(value)
    date_match = re.search(r"\d{2}[-]\d{2}[-]\d{4}", value_str)

    if date_match:
        date_str = date_match.group()
        try:
            arrival_date = pd.to_datetime(date_str, format="%d-%m-%Y")
            # arrival_date = date.strftime("%d-%m-%Y")
        except:
            arrival_date = pd.NaT
    else:
        arrival_date = pd.NaT

    if "REAL" in value_str.upper():
        arrival_type = "REAL"
    elif "REPROGRAMADO" in value_str.upper():
        arrival_type = "REPROGRAMADO"
    else:
        arrival_type = "PROYECTADO"

    return arrival_date, arrival_type


def get_previous_week(week_range, year=2024):
    # Extract the first week number
    match = re.search(r"W(\d+)", week_range)
    if not match:
        raise ValueError(f"Invalid week range format: {week_range}")

    week_num = int(match.group(1))

    # Create a date object for the Monday of the given week
    date = datetime.strptime(f"{year}-W{week_num}-1", "%Y-W%W-%w")

    # Subtract one week
    previous_week = date - timedelta(weeks=1)

    # Format the result
    return previous_week.strftime("%Y-W%V")


def format_datetime(value):
    if isinstance(value, pd.Timestamp) or isinstance(value, datetime):
        return value.strftime("%d-%m-%Y")
    elif isinstance(value, str):
        try:
            # Try to parse the string as a datetime
            parsed_date = pd.to_datetime(value, errors="raise")
            return parsed_date.strftime("%d-%m-%Y")
        except (ValueError, TypeError):
            # If parsing fails, return the original string
            return value
    else:
        return value


def read_component_arrivals():

    graph_client = GraphClient(acquire_token_func)
    site = graph_client.sites.get_by_url("https://globalkomatsu.sharepoint.com/sites/KCHCLSP00060")

    file = (
        site.drive.root.get_by_path(
            "/3.-%20Solicitudes%20Soporte%20Componentes/1.-%20Escondida/1.-%20Entrega%20de%20componentes/CONSOLIDADO%20DE%20PLANES/2024/Planilla%20de%20seguimiento%20de%20cumplimiento%20de%20entrega%20componentes%202024.xlsx"
        )
        .get()
        .execute_query()
    )
    content = file.get_content().execute_query().value
    data = BytesIO(content)
    workbook = openpyxl.load_workbook(data)
    frames = []
    sheetnames = [c for c in workbook.sheetnames if c not in ["W01", "W02", "W03", "W04", "W05"]]

    # sheetnames = workbook.sheetnames
    for sheet in sheetnames:
        # print(sheet)
        df = pd.read_excel(data, sheet_name=sheet, skiprows=1, engine="openpyxl", dtype="str")
        # Apply the formatting function to every cell in the DataFrame
        df = df.applymap(format_datetime)

        # Assume df is your DataFrame
        df = rename_datetime_columns(df)
        df = df[["Componente", "N°"] + [col for col in df.columns if re.match(r"^\d{4}-\d{2}-\d{2}$", col)]]
        # Rellenar los componentes hacia adelante para cubrir el número 1 y 2.
        # Por defecto la primera fila sólo debiese tener las semanas y el ffill no cambiará el hecho de que sea nula y se pueda sacar
        df = df.assign(
            Componente=df["Componente"].ffill(),
            arrival_projection_week=(
                get_previous_week(sheet, 2025)
                if sheet in ["W01", "W02", "W03", "W04", "W05"]
                else get_previous_week(sheet, 2024)
            ),
        )
        df = df.dropna(subset=["Componente", "N°"])

        df = df.pipe(reshape_to_long_format).dropna(subset=["value"])
        df[["arrival_date", "arrival_type"]] = df["value"].apply(lambda x: pd.Series(extract_date_and_type(x)))
        # Add an overall assertion check

        assert (
            df[df["value"].notna()]["arrival_date"].notna().all()
            and df[df["value"].notna()]["arrival_type"].notna().all()
        ), f"{sheet}: Some non-null values resulted in null arrival_date or arrival_type {df.loc[(df['value'].notna()) & (df['arrival_date'].isnull())].to_markdown()}"
        frames.append(df)

    # return frames
    df = pd.concat(frames)

    df = df.assign(
        component=df["Componente"].map(
            lambda x: {
                "Blower parrillas 930 - 960": "blower_parrilla",
                "Cil. de dirección 960": "cilindro_direccion",
                "Cil. de levante 960": "cilindro_levante",
                "Motor de tracción 960": "motor_traccion",
                "Módulo Potencia 960": "modulo_potencia",
                "Suspensión delantera 960": "suspension_delantera",
                "Suspensión trasera 960": "suspension_trasera",
            }.get(x)
        )
    ).dropna(subset=["component"])

    # Convert weeks to datetime for proper sorting
    df["arrival_projection_week"] = pd.to_datetime(df["arrival_projection_week"] + "-1", format="%Y-W%W-%w")
    df["arrival_week"] = pd.to_datetime(df["arrival_week"] + "-1", format="%Y-W%W-%w")
    # df["arrival_date"] = pd.to_datetime(df["arrival_date"], format="%d-%m-%Y")

    # Sort values to ensure the latest arrival_projection_week comes last
    df_latest = df.groupby(["arrival_week"])["arrival_projection_week"].max().reset_index()
    # Merge the original DataFrame with the latest arrival_projection_week DataFrame
    df = pd.merge(df, df_latest, on=["arrival_week", "arrival_projection_week"], how="inner")

    # Convert back to week format
    df["arrival_projection_week"] = df["arrival_projection_week"].dt.strftime("%Y-W%W")
    df["arrival_week"] = df["arrival_week"].dt.strftime("%Y-W%W")

    return df
